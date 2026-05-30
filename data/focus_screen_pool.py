"""Focused visuals for the consultant's screen pool only.

Strips out the 749 no_material (resolved) and the 224 clear-enforcement (also
resolved). Renders just the 19 Russell-1000 companies that ARE the deliverable
(reform_leaning + cross_exposure), with names directly on the chart and a
companion action-list workbook with all evidence in one place.

Outputs (in data/):
  screen_pool_quadrant.html / .png   — focused quadrant, just the 19, labeled
  screen_pool_action_list.xlsx       — Ticker / Company / Sector / stance / evidence URLs
"""
from __future__ import annotations

import math
import os
import sqlite3
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

os.chdir(Path(__file__).resolve().parent)
DB = "institutional_risk.db"
SCORED = "police_policy_scored_20260529.xlsx"


def fetch_screen_pool():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    rows = conn.execute("""
        SELECT i.ticker, i.company_name, i.sector,
               i.net_position, i.confidence,
               i.anti_police_type, i.anti_police_first_date,
               i.anti_police_current_status, i.anti_police_summary,
               i.anti_police_evidence_url,
               i.pro_police_type, i.pro_police_first_date,
               i.pro_police_current_status, i.pro_police_summary,
               i.pro_police_evidence_url,
               i.net_summary, i.investigation_model,
               u.market_value_musd
        FROM company_stance_investigation i
        JOIN universe u ON u.ticker = i.ticker
        WHERE i.net_position IN ('reform_leaning','cross_exposure')
        ORDER BY
            CASE i.net_position WHEN 'reform_leaning' THEN 0 ELSE 1 END,
            i.confidence DESC, i.ticker
    """).fetchall()
    conn.close()
    return rows


def render_quadrant(focus_df: pd.DataFrame, n: int):
    """Focused interactive quadrant: just the 19, large markers, ticker labels.

    Upper cluster shows labels next to markers. The dense pure-reform row at
    y~0 uses vertical "stem" leader lines staggered in height with the ticker
    label at the top of each stem, so every ticker is legible even when
    multiple markers share an x score.
    """
    cmap = {"Reform-aligned": "#2b8cbe", "Mixed engagement": "#756bb1"}
    top = focus_df[focus_df["Public-safety-alignment"] >= 20]
    bot = focus_df[focus_df["Public-safety-alignment"] < 20].sort_values(
        "Reform-alignment")

    def hovertext(sub):
        return sub.apply(
            lambda r: (
                f"<b>{r['Company']} ({r['Ticker']})</b><br>"
                f"Sector: {r['Sector']}<br>"
                f"Reform-alignment: {r['Reform-alignment']:.1f}<br>"
                f"Public-safety-alignment: {r['Public-safety-alignment']:.1f}<br>"
                f"Confidence: {r['Confidence']:.2f}<br>"
                f"Anti type: {r.get('Anti type','')}<br>"
                f"Pro type: {r.get('Pro type','')}"
            ), axis=1)

    fig = go.Figure()
    seen_in_legend: set[str] = set()
    # Upper cluster: labels next to markers (well-spread, no overlap issues).
    for cat, sub in top.groupby("Category"):
        fig.add_trace(go.Scatter(
            x=sub["Reform-alignment"], y=sub["Public-safety-alignment"],
            mode="markers+text",
            text=sub["Ticker"],
            textposition="top right",
            textfont=dict(size=14, color="#111", family="Arial Black"),
            marker=dict(size=22, color=cmap.get(cat, "#999"),
                        line=dict(width=1.2, color="white"),
                        opacity=0.92),
            name=cat,
            hovertext=hovertext(sub), hoverinfo="text",
        ))
        seen_in_legend.add(cat)
    # Bottom row: markers only — labels go into the right-side callout column.
    for cat, sub in bot.groupby("Category"):
        fig.add_trace(go.Scatter(
            x=sub["Reform-alignment"], y=sub["Public-safety-alignment"],
            mode="markers",
            marker=dict(size=22, color=cmap.get(cat, "#999"),
                        line=dict(width=1.2, color="white"),
                        opacity=0.92),
            name=cat, showlegend=(cat not in seen_in_legend),
            hovertext=hovertext(sub), hoverinfo="text",
        ))
        seen_in_legend.add(cat)
    fig.add_hline(y=50, line=dict(color="#ddd", width=1))
    fig.add_vline(x=50, line=dict(color="#ddd", width=1))
    qa = dict(showarrow=False, font=dict(size=11, color="#888"))
    fig.add_annotation(x=15, y=96, text="High public-safety, low reform", **qa)
    fig.add_annotation(x=85, y=96, text="Mixed (both sides)", **qa)
    fig.add_annotation(x=15, y=4, text="Limited on both", **qa)
    fig.add_annotation(x=85, y=4, text="Pure reform-aligned", **qa)
    # Vertical "stem" leaders for the dense bottom row: each marker gets a
    # vertical line going straight up, with the ticker label at the top.
    # Markers that share an x get small horizontal offsets so their stems and
    # labels separate; markers themselves stay at the true (x, y) score.
    if len(bot):
        bot_sorted = bot.sort_values(["Reform-alignment", "Ticker"]).reset_index(drop=True)
        stem_heights = np.linspace(40, 8, len(bot_sorted))
        offsets: list[float] = [0.0] * len(bot_sorted)
        rounded_x = bot_sorted["Reform-alignment"].round(1).tolist()
        i = 0
        while i < len(bot_sorted):
            j = i
            while j < len(bot_sorted) and rounded_x[j] == rounded_x[i]:
                j += 1
            grp = list(range(i, j))
            mid = (len(grp) - 1) / 2.0
            for k, idx in enumerate(grp):
                offsets[idx] = (k - mid) * 1.2
            i = j

        stem_x, stem_y = [], []
        label_x, label_y_, label_txt = [], [], []
        for idx, (_, r) in enumerate(bot_sorted.iterrows()):
            mx, my = r["Reform-alignment"], r["Public-safety-alignment"]
            lx = mx + offsets[idx]
            ly = stem_heights[idx]
            # leader line: marker → top, with a None pair to break between stems
            stem_x.extend([mx, lx, None])
            stem_y.extend([my, ly, None])
            label_x.append(lx)
            label_y_.append(ly)
            label_txt.append(r["Ticker"])

        fig.add_trace(go.Scatter(
            x=stem_x, y=stem_y, mode="lines",
            line=dict(color="rgba(120,120,120,0.7)", width=0.8),
            hoverinfo="skip", showlegend=False,
        ))
        fig.add_trace(go.Scatter(
            x=label_x, y=label_y_, mode="text", text=label_txt,
            textposition="top center",
            textfont=dict(size=13, color="#111", family="Arial Black"),
            hoverinfo="skip", showlegend=False,
        ))
    fig.update_layout(
        title=f"Screen Pool — Reform-aligned + Mixed-engagement ({n} Russell 1000 companies)",
        xaxis=dict(title="Reform-alignment score", range=[-5, 105]),
        yaxis=dict(title="Public-safety-alignment score", range=[-5, 105]),
        template="plotly_white", width=1500, height=1050,
        legend=dict(title="Category"),
    )
    fig.write_html("screen_pool_quadrant.html")

    fig2, ax = plt.subplots(figsize=(16, 11))
    for cat, sub in focus_df.groupby("Category"):
        ax.scatter(sub["Reform-alignment"], sub["Public-safety-alignment"],
                   s=320, c=cmap.get(cat, "#999"), alpha=0.92,
                   edgecolors="white", linewidths=1.2, label=cat)
    # Hybrid label layout: the upper cluster spreads with adjustText; the
    # dense pure-reform row at y=0 (10 labels in a narrow x band) is laid out
    # as a vertical column on the right with leader lines so each ticker is
    # legible. Markers stay at the true (x,y) scores.
    from adjustText import adjust_text
    top = focus_df[focus_df["Public-safety-alignment"] >= 20]
    bot = focus_df[focus_df["Public-safety-alignment"] < 20].sort_values(
        "Reform-alignment")

    if len(top):
        texts = [ax.text(r["Reform-alignment"], r["Public-safety-alignment"],
                         r["Ticker"], fontsize=13, color="#111",
                         fontweight="bold", zorder=5)
                 for _, r in top.iterrows()]
        adjust_text(
            texts, ax=ax,
            arrowprops=dict(arrowstyle="-", color="#888", lw=0.6, alpha=0.7),
            expand_points=(1.5, 1.6), expand_text=(1.2, 1.4),
            force_text=(0.6, 0.8), force_points=(0.4, 0.6),
            only_move={"text": "xy"},
        )

    if len(bot):
        col_x = 92.0
        ys = np.linspace(38, 2, len(bot))  # top-to-bottom matches sorted-by-x
        for (_, r), label_y in zip(bot.iterrows(), ys):
            ax.plot([r["Reform-alignment"], col_x - 1.2],
                    [r["Public-safety-alignment"], label_y],
                    color="#888", linewidth=0.5, alpha=0.7, zorder=1)
            ax.text(col_x, label_y, r["Ticker"], fontsize=12, color="#111",
                    fontweight="bold", ha="left", va="center", zorder=5)
    ax.axhline(50, color="#ddd"); ax.axvline(50, color="#ddd")
    for x, y, t in [(12, 96, "High PS, low reform"), (78, 96, "Mixed (both sides)"),
                    (12, 4, "Limited on both"), (78, 4, "Pure reform-aligned")]:
        ax.text(x, y, t, color="#999", fontsize=10)
    ax.set_xlim(-5, 105); ax.set_ylim(-5, 105)
    ax.set_xlabel("Reform-alignment score")
    ax.set_ylabel("Public-safety-alignment score")
    ax.set_title(f"Screen Pool — Reform-aligned + Mixed-engagement ({n} Russell 1000 companies)")
    ax.legend(title="Category", loc="lower left", framealpha=0.95)
    fig2.tight_layout()
    fig2.savefig("screen_pool_quadrant.png", dpi=160)


def render_action_list(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Screen pool action list"

    headers = [
        "Ticker", "Company", "Sector", "Mkt cap ($M)", "Net position",
        "Confidence", "Re-checked by Opus",
        "Reform action type", "Reform date", "Reform status",
        "Reform summary", "Reform evidence URL",
        "Enforcement action type", "Enforcement date", "Enforcement status",
        "Enforcement summary", "Enforcement evidence URL",
        "Net summary",
    ]
    ws.append(headers)
    hdr_fill = PatternFill("solid", fgColor="305AAA")
    hdr_font = Font(bold=True, color="FFFFFF")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for r in rows:
        rechecked = "yes" if r["investigation_model"] == "claude-opus-4-7" else "no"
        ws.append([
            r["ticker"], r["company_name"], r["sector"],
            round(r["market_value_musd"] or 0),
            r["net_position"], r["confidence"], rechecked,
            r["anti_police_type"], r["anti_police_first_date"], r["anti_police_current_status"],
            r["anti_police_summary"], r["anti_police_evidence_url"],
            r["pro_police_type"], r["pro_police_first_date"], r["pro_police_current_status"],
            r["pro_police_summary"], r["pro_police_evidence_url"],
            r["net_summary"],
        ])
        for c in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=c).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[ws.max_row].height = 90

    widths = {1: 8, 2: 26, 3: 22, 4: 12, 5: 18, 6: 10, 7: 10,
              8: 22, 9: 12, 10: 12, 11: 50, 12: 26,
              13: 24, 14: 12, 15: 12, 16: 50, 17: 26, 18: 60}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "C2"
    wb.save("screen_pool_action_list.xlsx")


def main():
    rows = fetch_screen_pool()
    n = len(rows)
    if n == 0:
        print("No screen-pool companies found."); return

    scored = pd.read_excel(SCORED, sheet_name="Company Scores")
    tickers = [r["ticker"] for r in rows]
    focus = scored[scored["Ticker"].isin(tickers)].copy()
    print(f"Screen pool: {n} companies (R1000, reform_leaning + cross_exposure)")
    print(focus[["Ticker", "Company", "Sector", "Category",
                 "Reform-alignment", "Public-safety-alignment", "Confidence"]]
          .to_string(index=False))

    render_quadrant(focus, n)
    render_action_list(rows)
    print("\nWrote:")
    print("  screen_pool_quadrant.html")
    print("  screen_pool_quadrant.png")
    print("  screen_pool_action_list.xlsx")


if __name__ == "__main__":
    main()
