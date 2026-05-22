"""
Add a "Score Charts" sheet to the workbook with:
  1. Histogram of policy_stance_score buckets (254 companies, all sectors)
  2. Top 25 enforcement-leaning (horizontal bar, red)
  3. Reform-leaning + cross-exposure (horizontal bar, blue)
  4. Sector mean-score chart (vertical bar)

Uses openpyxl BarChart natively — no external image generation.
"""
from __future__ import annotations

import sys
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config import DATA_DIR, RISK_DB_PATH
from src.risk_schema import connect


HEADER_FILL = PatternFill("solid", fgColor="333333")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True)
SECTION_FONT = Font(name="Calibri", size=12, bold=True)


def main() -> None:
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d")
    xlsx_path = Path(DATA_DIR) / f"police_policy_stance_{stamp}.xlsx"
    if not xlsx_path.exists():
        print(f"Workbook not found: {xlsx_path}")
        return

    conn = connect(RISK_DB_PATH)
    wb = load_workbook(xlsx_path)

    if "Score Charts" in wb.sheetnames:
        del wb["Score Charts"]
    ws = wb.create_sheet("Score Charts", 2)

    ws["A1"] = "Policy Stance Scores — Visualizations"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:H1")
    ws["A2"] = ("Score scale: +5 strong enforcement-leaning ... 0 neutral ... -5 strong reform-leaning. "
                "Derived from anti/pro type weights × current_status modifier + fed LE contract bonus.")
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:H2")

    # ── Section 1: Distribution histogram across all 254 companies ───────
    ws["A4"] = "1. Score distribution across all 254 investigated companies"
    ws["A4"].font = SECTION_FONT

    buckets = [
        ("+3 to +5", "Strong enforcement-leaning", "FFC7CE"),
        ("+1 to +3", "Lean enforcement", "F4CCCC"),
        ("-1 to +1", "No material posture",   "EFEFEF"),
        ("-1 to -3", "Lean reform",            "DCE7F2"),
        ("-3 to -5", "Strong reform-leaning", "9DC3E6"),
    ]
    bucket_counts = {b[0]: 0 for b in buckets}
    for r in conn.execute("SELECT policy_stance_score FROM company_stance_investigation WHERE policy_stance_score IS NOT NULL;"):
        s = r["policy_stance_score"]
        if   s >=  3: bucket_counts["+3 to +5"] += 1
        elif s >=  1: bucket_counts["+1 to +3"] += 1
        elif s >  -1: bucket_counts["-1 to +1"] += 1
        elif s >  -3: bucket_counts["-1 to -3"] += 1
        else:         bucket_counts["-3 to -5"] += 1

    ws["A6"] = "Bucket"
    ws["B6"] = "Label"
    ws["C6"] = "Count"
    for c in range(1, 4):
        cell = ws.cell(row=6, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, (key, label, color) in enumerate(buckets, start=7):
        ws.cell(row=i, column=1, value=key)
        ws.cell(row=i, column=2, value=label)
        ws.cell(row=i, column=3, value=bucket_counts[key])
        ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor=color)
        ws.cell(row=i, column=2).fill = PatternFill("solid", fgColor=color)

    # Build vertical bar chart
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 11
    chart1.title = "Companies by score bucket (n=254)"
    chart1.x_axis.title = "Score bucket (+ = enforcement, - = reform)"
    chart1.y_axis.title = "Number of companies"
    data = Reference(ws, min_col=3, min_row=6, max_row=11, max_col=3)
    cats = Reference(ws, min_col=2, min_row=7, max_row=11)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.dataLabels = DataLabelList(showVal=True)
    chart1.height = 10
    chart1.width = 18
    ws.add_chart(chart1, "E4")

    # ── Section 2: Top 25 enforcement-leaning ───────────────────────────
    base_row = 30
    ws.cell(row=base_row, column=1, value="2. Top 25 enforcement-leaning tickers").font = SECTION_FONT

    top_enforce = conn.execute("""
        SELECT ticker, company_name, policy_stance_score
        FROM company_stance_investigation
        WHERE policy_stance_score IS NOT NULL
        ORDER BY policy_stance_score DESC, ticker
        LIMIT 25;
    """).fetchall()

    hdr_row = base_row + 2
    ws.cell(row=hdr_row, column=1, value="Ticker")
    ws.cell(row=hdr_row, column=2, value="Company")
    ws.cell(row=hdr_row, column=3, value="Score")
    for c in range(1, 4):
        cell = ws.cell(row=hdr_row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for i, r in enumerate(top_enforce, start=hdr_row + 1):
        ws.cell(row=i, column=1, value=r["ticker"])
        ws.cell(row=i, column=2, value=r["company_name"])
        ws.cell(row=i, column=3, value=r["policy_stance_score"])
        ws.cell(row=i, column=3).number_format = '+#,##0.00;-#,##0.00;0.00'

    chart2 = BarChart()
    chart2.type = "bar"     # horizontal
    chart2.style = 11
    chart2.title = "Top 25 enforcement-leaning (highest +scores)"
    chart2.x_axis.title = "Policy stance score (+ = enforcement)"
    data = Reference(ws, min_col=3, min_row=hdr_row, max_row=hdr_row + 25)
    cats = Reference(ws, min_col=1, min_row=hdr_row + 1, max_row=hdr_row + 25)
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.height = 16
    chart2.width = 22
    ws.add_chart(chart2, f"E{base_row}")

    # ── Section 3: Reform-leaning + cross-exposure ──────────────────────
    base_row3 = base_row + 60
    ws.cell(row=base_row3, column=1,
            value="3. Reform-leaning + cross-exposure (negative + near-zero scores)").font = SECTION_FONT

    reform_etc = conn.execute("""
        SELECT ticker, company_name, policy_stance_score
        FROM company_stance_investigation
        WHERE policy_stance_score IS NOT NULL
          AND policy_stance_score < 1.5
          AND net_position IN ('reform_leaning', 'cross_exposure', 'no_material_exposure')
          AND (anti_police_action = 1 OR pro_police_action = 1)
        ORDER BY policy_stance_score ASC, ticker
        LIMIT 25;
    """).fetchall()

    hdr_row3 = base_row3 + 2
    ws.cell(row=hdr_row3, column=1, value="Ticker")
    ws.cell(row=hdr_row3, column=2, value="Company")
    ws.cell(row=hdr_row3, column=3, value="Score")
    for c in range(1, 4):
        cell = ws.cell(row=hdr_row3, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    n3 = 0
    for i, r in enumerate(reform_etc, start=hdr_row3 + 1):
        ws.cell(row=i, column=1, value=r["ticker"])
        ws.cell(row=i, column=2, value=r["company_name"])
        ws.cell(row=i, column=3, value=r["policy_stance_score"])
        ws.cell(row=i, column=3).number_format = '+#,##0.00;-#,##0.00;0.00'
        n3 += 1

    chart3 = BarChart()
    chart3.type = "bar"
    chart3.style = 12
    chart3.title = "Reform-leaning + cross-exposure (negative + near-zero)"
    chart3.x_axis.title = "Policy stance score (- = reform)"
    data = Reference(ws, min_col=3, min_row=hdr_row3, max_row=hdr_row3 + n3)
    cats = Reference(ws, min_col=1, min_row=hdr_row3 + 1, max_row=hdr_row3 + n3)
    chart3.add_data(data, titles_from_data=True)
    chart3.set_categories(cats)
    chart3.height = 12
    chart3.width = 22
    ws.add_chart(chart3, f"E{base_row3}")

    # ── Section 4: Mean score by sector ─────────────────────────────────
    base_row4 = base_row3 + 60
    ws.cell(row=base_row4, column=1, value="4. Average score by sector").font = SECTION_FONT

    sector_rows = conn.execute("""
        SELECT sector, AVG(policy_stance_score) AS mean_score,
               COUNT(*) AS n
        FROM company_stance_investigation
        WHERE sector IS NOT NULL AND sector != ''
          AND policy_stance_score IS NOT NULL
        GROUP BY sector
        ORDER BY mean_score DESC;
    """).fetchall()

    hdr_row4 = base_row4 + 2
    ws.cell(row=hdr_row4, column=1, value="Sector")
    ws.cell(row=hdr_row4, column=2, value="N")
    ws.cell(row=hdr_row4, column=3, value="Mean score")
    for c in range(1, 4):
        cell = ws.cell(row=hdr_row4, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for i, r in enumerate(sector_rows, start=hdr_row4 + 1):
        ws.cell(row=i, column=1, value=r["sector"])
        ws.cell(row=i, column=2, value=r["n"])
        ws.cell(row=i, column=3, value=round(r["mean_score"], 2))
        ws.cell(row=i, column=3).number_format = '+#,##0.00;-#,##0.00;0.00'

    chart4 = BarChart()
    chart4.type = "col"
    chart4.style = 11
    chart4.title = "Average policy_stance_score by sector"
    chart4.x_axis.title = "Sector"
    chart4.y_axis.title = "Mean score (+ enforcement / - reform)"
    data = Reference(ws, min_col=3, min_row=hdr_row4, max_row=hdr_row4 + len(sector_rows))
    cats = Reference(ws, min_col=1, min_row=hdr_row4 + 1, max_row=hdr_row4 + len(sector_rows))
    chart4.add_data(data, titles_from_data=True)
    chart4.set_categories(cats)
    chart4.height = 12
    chart4.width = 22
    ws.add_chart(chart4, f"E{base_row4}")

    # Column widths
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 11

    wb.save(xlsx_path)
    print(f"Added Score Charts sheet -> {xlsx_path}")
    print(f"Workbook now has {len(wb.sheetnames)} sheets")
    conn.close()


if __name__ == "__main__":
    main()
