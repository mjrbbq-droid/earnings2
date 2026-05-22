"""
Per-company dossier — every signal from every source joined on ticker.

For each ticker:
  - v3 investigation: anti+pro types, dates, status, net_position
  - Foundation 990: trajectory + Schedule I anti-adjacent grants
  - DAFs they fund (from Schedule I 'daf_intermediary' grants)
  - Federal LE contracts (USAspending): n + total $
  - SEC EDGAR keyword hits
  - GDELT news count

Writes to a new sheet "Company Dossier" in the Excel workbook.
"""
from __future__ import annotations

import sys
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config import DATA_DIR, RISK_DB_PATH
from src.risk_schema import connect


NET_FILLS = {
    "reform_leaning":       "DCE7F2",
    "enforcement_leaning":  "F4CCCC",
    "cross_exposure":       "FCE5CD",
    "no_material_exposure": "EFEFEF",
    "unknown":              "FFFFFF",
}
HEADER_FILL = PatternFill("solid", fgColor="333333")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")


def fetch_dossier(conn) -> list[dict]:
    rows = conn.execute(
        """
        SELECT
            csi.ticker, csi.company_name, csi.sector,
            csi.net_position, csi.policy_stance_score,
            csi.net_summary, csi.confidence,
            csi.anti_police_action, csi.anti_police_type,
            csi.anti_police_first_date, csi.anti_police_current_status,
            csi.anti_police_summary,
            csi.pro_police_action, csi.pro_police_type,
            csi.pro_police_first_date, csi.pro_police_current_status,
            csi.pro_police_summary,

            df.ein                          AS foundation_ein,
            df.foundation_name              AS foundation_name,

            -- 990 totals + classified grant rollups
            COALESCE(ff_22.total_grants_paid, 0) AS grants_paid_2022,
            COALESCE(ff_23.total_grants_paid, 0) AS grants_paid_2023,

            (SELECT COALESCE(SUM(grant_amount), 0)
             FROM foundation_grants
             WHERE donor_ein = df.ein
               AND grantee_classification = 'anti_police_adversarial')
                AS direct_anti_police_total,

            (SELECT COALESCE(SUM(grant_amount), 0)
             FROM foundation_grants
             WHERE donor_ein = df.ein
               AND grantee_classification IN ('broad_cj_reform', 'collaborative_reform', 'reentry_employment'))
                AS direct_cj_reform_total,

            (SELECT COALESCE(SUM(grant_amount), 0)
             FROM foundation_grants
             WHERE donor_ein = df.ein
               AND grantee_classification = 'pro_police')
                AS direct_pro_police_total,

            (SELECT COALESCE(SUM(grant_amount), 0)
             FROM foundation_grants
             WHERE donor_ein = df.ein
               AND grantee_classification = 'daf_intermediary')
                AS daf_outflow_total,

            -- Federal LE contracts
            (SELECT COUNT(*) FROM company_federal_contracts WHERE ticker = csi.ticker AND is_le_agency = 1) AS n_fed_le_contracts,
            (SELECT COALESCE(SUM(award_amount), 0) FROM company_federal_contracts WHERE ticker = csi.ticker AND is_le_agency = 1) AS fed_le_total,

            -- SEC EDGAR signals
            (SELECT COUNT(*) FROM company_sec_signals WHERE ticker = csi.ticker) AS sec_hits,

            -- GDELT news
            (SELECT COUNT(*) FROM gdelt_news WHERE ticker = csi.ticker) AS gdelt_count

        FROM company_stance_investigation csi
        LEFT JOIN donor_foundations df
               ON df.donor_ticker = csi.ticker AND df.relationship != 'daf'
        LEFT JOIN foundation_filings ff_22
               ON ff_22.ein = df.ein AND ff_22.tax_year = 2022
        LEFT JOIN foundation_filings ff_23
               ON ff_23.ein = df.ein AND ff_23.tax_year = 2023
        WHERE csi.net_position IN ('reform_leaning', 'cross_exposure', 'enforcement_leaning')
        ORDER BY
            csi.policy_stance_score ASC,    -- most reform-leaning first
            csi.ticker;
        """
    ).fetchall()
    return [dict(r) for r in rows]


def main() -> None:
    conn = connect(RISK_DB_PATH)
    rows = fetch_dossier(conn)
    print(f"Dossier rows: {len(rows)}")

    xlsx_path = Path(DATA_DIR) / f"police_policy_stance_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    if not xlsx_path.exists():
        print(f"Excel workbook not found: {xlsx_path}")
        return

    wb = load_workbook(xlsx_path)
    if "Company Dossier" in wb.sheetnames:
        del wb["Company Dossier"]
    ws = wb.create_sheet("Company Dossier", 1)

    headers = [
        "Ticker", "Company", "Sector", "Net Position",
        "Policy stance score",
        "Reform-side type", "First date", "Status",
        "Enforcement-side type", "First date", "Status",
        "Foundation",
        "Grants paid FY22 ($M)", "Grants paid FY23 ($M)",
        "Direct reform-advocacy $", "Direct CJ-reform $", "Direct police-foundation $",
        "DAF outflow $ (opaque layer)",
        "Fed LE contracts (n)", "Fed LE total $",
        "SEC EDGAR hits",
        "Stance summary",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.freeze_panes = "D2"
    ws.row_dimensions[1].height = 36

    for r in rows:
        net = r["net_position"] or "unknown"
        fill_color = NET_FILLS.get(net, "FFFFFF")
        ws.append([
            r["ticker"], r["company_name"], r["sector"], net,
            r["policy_stance_score"],
            r["anti_police_type"] if r["anti_police_action"] else "",
            r["anti_police_first_date"] or "",
            r["anti_police_current_status"] if r["anti_police_action"] else "",
            r["pro_police_type"] if r["pro_police_action"] else "",
            r["pro_police_first_date"] or "",
            r["pro_police_current_status"] if r["pro_police_action"] else "",
            r["foundation_name"] or "—",
            (r["grants_paid_2022"] / 1e6) if r["grants_paid_2022"] else None,
            (r["grants_paid_2023"] / 1e6) if r["grants_paid_2023"] else None,
            r["direct_anti_police_total"] or 0,
            r["direct_cj_reform_total"] or 0,
            r["direct_pro_police_total"] or 0,
            r["daf_outflow_total"] or 0,
            r["n_fed_le_contracts"] or 0,
            r["fed_le_total"] or 0,
            r["sec_hits"] or 0,
            r["net_summary"] or "",
        ])
        row_idx = ws.max_row
        fill = PatternFill("solid", fgColor=fill_color)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = fill
            cell.alignment = WRAP

        # Number formats
        # Col 5 = policy_stance_score
        ws.cell(row=row_idx, column=5).number_format = '+#,##0.00;-#,##0.00;0.00'
        # Cols 13, 14 = grants paid FY22/23 (already in $M)
        for c in (13, 14):
            ws.cell(row=row_idx, column=c).number_format = '"$"#,##0.00"M"'
        # Cols 15-18 = direct $ amounts, 20 = fed LE total
        for c in (15, 16, 17, 18, 20):
            ws.cell(row=row_idx, column=c).number_format = '"$"#,##0'

        # Bold the score cell + color by sign
        score = r["policy_stance_score"]
        if score is not None:
            sc_cell = ws.cell(row=row_idx, column=5)
            sc_cell.font = Font(bold=True, color=(
                "9C0006" if score >= 2 else ("003366" if score <= -2 else "333333")
            ))

        ws.row_dimensions[row_idx].height = 75

    widths = {
        1: 7, 2: 26, 3: 22, 4: 22, 5: 11,
        6: 32, 7: 12, 8: 14,
        9: 32, 10: 12, 11: 14,
        12: 30,
        13: 14, 14: 14,
        15: 16, 16: 16, 17: 16, 18: 18,
        19: 11, 20: 16,
        21: 10,
        22: 60,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    wb.save(xlsx_path)
    print(f"Saved Company Dossier sheet -> {xlsx_path}")
    conn.close()


if __name__ == "__main__":
    main()
