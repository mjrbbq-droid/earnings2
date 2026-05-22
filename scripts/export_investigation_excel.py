"""
Build a multi-sheet Excel review workbook from company_stance_investigation.

Sheets:
  1. Summary             — top-level counts, distribution charts (text)
  2. All Companies       — every row, anti+pro side-by-side
  3. Anti-police only    — companies with anti_police_action=1
  4. Pro-police only     — companies with pro_police_action=1
  5. Mixed               — net_position='mixed'
  6. Anti-police-net     — clean anti, no offsetting pro
  7. Pro-police-net      — clean pro, no offsetting anti
  8. Police-foundation funders   — pro_police_type='donates_to_police_foundations'

Formatting:
  - Frozen top row
  - Color-coded rows by net_position (anti=blue, pro=red, mixed=amber, neutral=gray)
  - Hyperlinked evidence URLs
  - Wide columns for summary/text fields
"""
from __future__ import annotations

import sys
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config import DATA_DIR, RISK_DB_PATH
from src.risk_schema import connect


NET_FILLS = {
    "reform_leaning":       PatternFill("solid", fgColor="DCE7F2"),   # soft blue
    "enforcement_leaning":  PatternFill("solid", fgColor="F4CCCC"),   # soft red
    "cross_exposure":       PatternFill("solid", fgColor="FCE5CD"),   # soft amber
    "no_material_exposure": PatternFill("solid", fgColor="EFEFEF"),   # soft gray
    "unknown":              PatternFill("solid", fgColor="FFFFFF"),
}
HEADER_FILL = PatternFill("solid", fgColor="333333")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT  = Font(name="Calibri", size=14, bold=True)
LINK_FONT   = Font(name="Calibri", size=10, color="0563C1", underline="single")
DEFAULT_FONT = Font(name="Calibri", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP  = Alignment(vertical="top")


def fetch_rows(conn) -> list[dict]:
    rows = conn.execute(
        """
        SELECT
            ticker, company_name, sector, net_position, net_summary,
            confidence, investigated_at_utc,
            anti_police_action, anti_police_type, anti_police_first_date,
            anti_police_first_year, anti_police_last_known_date,
            anti_police_current_status, anti_police_summary,
            anti_police_evidence_url, anti_police_evidence_quote,
            pro_police_action, pro_police_type, pro_police_first_date,
            pro_police_first_year, pro_police_last_known_date,
            pro_police_current_status, pro_police_summary,
            pro_police_evidence_url, pro_police_evidence_quote,
            notes
        FROM company_stance_investigation
        ORDER BY
            CASE net_position
                WHEN 'reform_leaning' THEN 1
                WHEN 'cross_exposure'  THEN 2
                WHEN 'enforcement_leaning'  THEN 3
                WHEN 'neutral'         THEN 4
                ELSE 5 END,
            COALESCE(anti_police_first_date, pro_police_first_date, '9999'),
            ticker;
        """
    ).fetchall()
    return [dict(r) for r in rows]


def _style_header(ws, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28


def _set_widths(ws, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _link(cell, url: str | None, label: str | None = None) -> None:
    if url:
        cell.value = label or url
        cell.hyperlink = url
        cell.font = LINK_FONT


def write_summary(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Summary", 0)
    ws["A1"] = "Anti-Police Stance Investigation — Summary"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws["A2"] = f"Generated {ts}  ·  S&P 500 priority sectors + extras  ·  Model: Claude Opus 4.7 with web_search"
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:F2")

    # Net position breakdown
    from collections import Counter
    net_counts = Counter(r["net_position"] for r in rows)
    total = len(rows)

    ws["A4"] = "Net position breakdown"
    ws["A4"].font = Font(bold=True, size=12)
    ws["A5"], ws["B5"], ws["C5"] = "Net position", "Count", "% of universe"
    for c in "ABC":
        ws[f"{c}5"].fill = HEADER_FILL
        ws[f"{c}5"].font = HEADER_FONT
    row = 6
    order = ["reform_leaning", "cross_exposure", "enforcement_leaning", "no_material_exposure", "unknown"]
    for k in order:
        if net_counts.get(k):
            ws.cell(row=row, column=1, value=k).fill = NET_FILLS.get(k, PatternFill())
            ws.cell(row=row, column=2, value=net_counts[k])
            ws.cell(row=row, column=3, value=f"{100 * net_counts[k] / total:.1f}%")
            row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=total).font = Font(bold=True)

    # Anti-police type breakdown
    row += 3
    ws.cell(row=row, column=1, value="Reform-side action types").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="Type"); ws.cell(row=row, column=2, value="Count")
    ws.cell(row=row, column=1).fill = HEADER_FILL; ws.cell(row=row, column=1).font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL; ws.cell(row=row, column=2).font = HEADER_FONT
    anti_counts = Counter(r["anti_police_type"] for r in rows if r["anti_police_action"])
    for t, n in anti_counts.most_common():
        row += 1
        ws.cell(row=row, column=1, value=t)
        ws.cell(row=row, column=2, value=n)

    # Pro-police type breakdown
    row += 3
    ws.cell(row=row, column=1, value="Enforcement-side action types").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="Type"); ws.cell(row=row, column=2, value="Count")
    ws.cell(row=row, column=1).fill = HEADER_FILL; ws.cell(row=row, column=1).font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL; ws.cell(row=row, column=2).font = HEADER_FONT
    pro_counts = Counter(r["pro_police_type"] for r in rows if r["pro_police_action"])
    for t, n in pro_counts.most_common():
        row += 1
        ws.cell(row=row, column=1, value=t)
        ws.cell(row=row, column=2, value=n)

    # Sheet legend
    row += 3
    ws.cell(row=row, column=1, value="Sheet legend").font = Font(bold=True, size=12)
    row += 1
    for sheet, desc in [
        ("All Companies",          "Master table — every row with anti+pro fields side-by-side"),
        ("Reform-leaning",         "Clean reform-side signal (no offsetting enforcement exposure)"),
        ("Enforcement-leaning",    "Clean enforcement-side exposure (no offsetting reform)"),
        ("Cross-exposure",         "Both reform and enforcement actions present — most analytically interesting"),
        ("Reform-side actions",    "Every company with any reform action (full detail)"),
        ("Enforcement-side actions", "Every company with any enforcement action (full detail)"),
        ("Police-foundation funders", "Companies donating to police foundations specifically"),
    ]:
        ws.cell(row=row, column=1, value=sheet).font = Font(bold=True)
        ws.cell(row=row, column=2, value=desc)
        row += 1

    _set_widths(ws, {1: 32, 2: 70, 3: 16})


MASTER_HEADERS = [
    "Ticker", "Company", "Sector", "Net Position", "Net Summary",
    "Anti?", "Anti type", "Anti first date", "Anti current status",
    "Anti summary", "Anti evidence URL",
    "Pro?", "Pro type", "Pro first date", "Pro current status",
    "Pro summary", "Pro evidence URL",
    "Confidence", "Notes",
]

MASTER_WIDTHS = {
    1: 8,  2: 30, 3: 22, 4: 18, 5: 60,
    6: 6,  7: 32, 8: 14, 9: 16,
    10: 60, 11: 38,
    12: 6,  13: 36, 14: 14, 15: 16,
    16: 60, 17: 38,
    18: 11, 19: 50,
}


def write_master(wb: Workbook, rows: list[dict], *, sheet_name: str, filter_fn=None) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.append(MASTER_HEADERS)
    _style_header(ws, len(MASTER_HEADERS))

    filtered = [r for r in rows if (filter_fn is None or filter_fn(r))]

    for r in filtered:
        net = r["net_position"] or "unknown"
        fill = NET_FILLS.get(net, PatternFill())
        ws.append([
            r["ticker"],
            r["company_name"],
            r["sector"],
            net,
            r["net_summary"],
            "Y" if r["anti_police_action"] else "",
            r["anti_police_type"] if r["anti_police_action"] else "",
            r["anti_police_first_date"] or "",
            r["anti_police_current_status"] if r["anti_police_action"] else "",
            r["anti_police_summary"] or "",
            "",  # URL — set below as hyperlink
            "Y" if r["pro_police_action"] else "",
            r["pro_police_type"] if r["pro_police_action"] else "",
            r["pro_police_first_date"] or "",
            r["pro_police_current_status"] if r["pro_police_action"] else "",
            r["pro_police_summary"] or "",
            "",  # URL — set below as hyperlink
            r["confidence"],
            r["notes"] or "",
        ])
        row_idx = ws.max_row

        # Color the row
        for c in range(1, len(MASTER_HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=c)
            if cell.font.color is None:
                cell.font = DEFAULT_FONT
            cell.fill = fill
            cell.alignment = WRAP

        # Hyperlinks
        if r["anti_police_evidence_url"]:
            _link(ws.cell(row=row_idx, column=11), r["anti_police_evidence_url"], "source")
        if r["pro_police_evidence_url"]:
            _link(ws.cell(row=row_idx, column=17), r["pro_police_evidence_url"], "source")

        ws.row_dimensions[row_idx].height = 80

    _set_widths(ws, MASTER_WIDTHS)


def write_focused_anti(wb: Workbook, rows: list[dict]) -> None:
    """Reform-side-only view: 11 columns, narrower."""
    ws = wb.create_sheet("Reform-side actions")
    headers = [
        "Ticker", "Company", "Sector", "First action date", "Anti type",
        "Current status", "Net position", "Summary", "Evidence URL",
        "Evidence quote", "Confidence",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    filtered = [r for r in rows if r["anti_police_action"]]
    filtered.sort(key=lambda r: (r["anti_police_first_date"] or "9999", r["ticker"]))

    for r in filtered:
        net = r["net_position"] or "unknown"
        fill = NET_FILLS.get(net, PatternFill())
        ws.append([
            r["ticker"],
            r["company_name"],
            r["sector"],
            r["anti_police_first_date"] or "",
            r["anti_police_type"] or "",
            r["anti_police_current_status"] or "",
            net,
            r["anti_police_summary"] or "",
            "",
            r["anti_police_evidence_quote"] or "",
            r["confidence"],
        ])
        row_idx = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = fill
            cell.alignment = WRAP
            cell.font = DEFAULT_FONT
        if r["anti_police_evidence_url"]:
            _link(ws.cell(row=row_idx, column=9), r["anti_police_evidence_url"], "source")
        ws.row_dimensions[row_idx].height = 100

    _set_widths(ws, {1: 8, 2: 28, 3: 22, 4: 14, 5: 32, 6: 14,
                     7: 18, 8: 70, 9: 12, 10: 60, 11: 11})


def write_focused_pro(wb: Workbook, rows: list[dict]) -> None:
    """Enforcement-side-only view."""
    ws = wb.create_sheet("Enforcement-side actions")
    headers = [
        "Ticker", "Company", "Sector", "First action date", "Pro type",
        "Current status", "Net position", "Summary", "Evidence URL",
        "Evidence quote", "Confidence",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    filtered = [r for r in rows if r["pro_police_action"]]
    filtered.sort(key=lambda r: (r["pro_police_first_date"] or "9999", r["ticker"]))

    for r in filtered:
        net = r["net_position"] or "unknown"
        fill = NET_FILLS.get(net, PatternFill())
        ws.append([
            r["ticker"],
            r["company_name"],
            r["sector"],
            r["pro_police_first_date"] or "",
            r["pro_police_type"] or "",
            r["pro_police_current_status"] or "",
            net,
            r["pro_police_summary"] or "",
            "",
            r["pro_police_evidence_quote"] or "",
            r["confidence"],
        ])
        row_idx = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = fill
            cell.alignment = WRAP
            cell.font = DEFAULT_FONT
        if r["pro_police_evidence_url"]:
            _link(ws.cell(row=row_idx, column=9), r["pro_police_evidence_url"], "source")
        ws.row_dimensions[row_idx].height = 100

    _set_widths(ws, {1: 8, 2: 28, 3: 22, 4: 14, 5: 32, 6: 14,
                     7: 18, 8: 70, 9: 12, 10: 60, 11: 11})


def write_recategorization_diff(wb: Workbook, rows: list[dict], snapshot_path: Path) -> None:
    """Compare v2 (pre-refinement) snapshot vs v3 (current) rows for the 30
    previously-anti-flagged tickers."""
    if not snapshot_path.exists():
        return
    v2_rows = {r["ticker"]: r for r in json.loads(snapshot_path.read_text())}
    v3_by_ticker = {r["ticker"]: r for r in rows}

    ws = wb.create_sheet("Recategorization v2 vs v3", 1)
    headers = [
        "Ticker", "Company",
        "v2 anti?", "v2 anti type", "v2 status", "v2 net",
        "v3 anti?", "v3 anti type", "v3 status", "v3 net",
        "Verdict", "v3 net_summary",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    moved_out = []
    still_anti = []
    for t, v2 in v2_rows.items():
        v3 = v3_by_ticker.get(t)
        if not v3:
            continue
        v2_anti = bool(v2["anti_police_action"])
        v3_anti = bool(v3["anti_police_action"])
        if v2_anti and not v3_anti:
            verdict = "MOVED OUT OF ANTI"
            moved_out.append(t)
        elif v2_anti and v3_anti:
            verdict = "still anti"
            still_anti.append(t)
        elif not v2_anti and v3_anti:
            verdict = "NEWLY anti"
        else:
            verdict = "no change"

        ws.append([
            t, v3["company_name"] or v2.get("company_name", ""),
            "Y" if v2_anti else "",
            v2.get("anti_police_type") or "",
            v2.get("anti_police_current_status") or "",
            v2.get("net_position") or "",
            "Y" if v3_anti else "",
            v3.get("anti_police_type") or "",
            v3.get("anti_police_current_status") or "",
            v3.get("net_position") or "",
            verdict,
            v3.get("net_summary") or "",
        ])
        row_idx = ws.max_row
        # Color the verdict cell
        v_cell = ws.cell(row=row_idx, column=11)
        if verdict == "MOVED OUT OF ANTI":
            v_cell.fill = PatternFill("solid", fgColor="C6EFCE")  # green
            v_cell.font = Font(bold=True, color="006100")
        elif verdict == "still anti":
            v_cell.fill = PatternFill("solid", fgColor="DCE7F2")  # soft blue
            v_cell.font = Font(bold=True, color="003366")
        elif verdict == "NEWLY anti":
            v_cell.fill = PatternFill("solid", fgColor="FFC7CE")  # red
            v_cell.font = Font(bold=True, color="9C0006")
        # net cells colored too
        for col_idx in (6, 10):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = NET_FILLS.get(cell.value, PatternFill())
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=c)
            if cell.alignment is None:
                cell.alignment = WRAP
        ws.row_dimensions[row_idx].height = 70

    _set_widths(ws, {1: 8, 2: 28, 3: 8, 4: 32, 5: 12, 6: 18,
                     7: 8, 8: 38, 9: 12, 10: 18, 11: 22, 12: 70})

    print(f"  Diff: {len(moved_out)} moved out of anti, {len(still_anti)} still anti")


def write_990_trajectory(wb: Workbook, conn) -> None:
    """Foundation total-grants-paid per year, from actual 990 filings."""
    rows = conn.execute(
        """
        SELECT d.donor_ticker, d.foundation_name, d.donor_company, d.relationship,
               d.notes, f.tax_year, f.total_grants_paid, f.form_type, f.pdf_url
        FROM foundation_filings f
        JOIN donor_foundations d ON d.ein = f.ein
        WHERE f.tax_year BETWEEN 2018 AND 2024
        ORDER BY d.donor_ticker, f.tax_year;
        """
    ).fetchall()

    if not rows:
        return

    # Pivot: ticker -> {year: grants}
    pivot: dict[str, dict[int, int | None]] = {}
    meta: dict[str, dict] = {}
    years_seen: set[int] = set()
    for r in rows:
        t = r["donor_ticker"]
        pivot.setdefault(t, {})[r["tax_year"]] = r["total_grants_paid"]
        meta.setdefault(t, {
            "foundation_name": r["foundation_name"],
            "donor_company":   r["donor_company"],
            "relationship":    r["relationship"],
            "notes":           r["notes"],
            "pdf_url":         r["pdf_url"],
        })
        years_seen.add(r["tax_year"])
    years = sorted(years_seen)

    ws = wb.create_sheet("990 Trajectory", 2)
    headers = ["Ticker", "Foundation", "Company"] + [str(y) for y in years] + ["Δ 2020→2023", "Verdict", "Notes"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for t in sorted(pivot.keys()):
        m = meta[t]
        row_data = [t, m["foundation_name"], m["donor_company"]]
        year_values: list[int | None] = []
        for y in years:
            g = pivot[t].get(y)
            year_values.append(g)
            row_data.append(g / 1e6 if g else None)

        # Compute Δ 2020→2023
        g2020 = pivot[t].get(2020)
        g2023 = pivot[t].get(2023)
        if g2020 and g2023:
            delta = (g2023 - g2020) / g2020
            row_data.append(f"{delta:+.0%}")
            if delta >= 0.15:
                verdict = "GROWING"
            elif delta <= -0.20:
                verdict = "SHRINKING"
            else:
                verdict = "flat"
        else:
            row_data.append("")
            verdict = "incomplete data"
        row_data.append(verdict)
        row_data.append(m["notes"] or "")
        ws.append(row_data)
        row_idx = ws.max_row

        # Style: numbers as currency-style, verdict colored
        for c in range(4, 4 + len(years)):
            cell = ws.cell(row=row_idx, column=c)
            cell.number_format = '"$"#,##0.00"M"'
            cell.alignment = Alignment(horizontal="right")

        v_cell = ws.cell(row=row_idx, column=len(years) + 5)
        if verdict == "GROWING":
            v_cell.fill = PatternFill("solid", fgColor="FFC7CE")  # pink — growing donor pool
            v_cell.font = Font(bold=True, color="9C0006")
        elif verdict == "SHRINKING":
            v_cell.fill = PatternFill("solid", fgColor="C6EFCE")  # green — shrinking
            v_cell.font = Font(bold=True, color="006100")
        elif verdict == "flat":
            v_cell.fill = PatternFill("solid", fgColor="EFEFEF")
        elif verdict == "incomplete data":
            v_cell.fill = PatternFill("solid", fgColor="FFEB9C")
            v_cell.font = Font(color="9C5700")

        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=c)
            if cell.alignment.wrap_text is False and c != len(years) + 5 and c <= 3:
                cell.alignment = TOP
            elif c == len(headers):
                cell.alignment = WRAP
        ws.row_dimensions[row_idx].height = 40

    widths = {1: 8, 2: 40, 3: 22}
    for i, y in enumerate(years):
        widths[4 + i] = 12
    widths[4 + len(years)]   = 12  # delta
    widths[4 + len(years)+1] = 16  # verdict
    widths[4 + len(years)+2] = 50  # notes
    _set_widths(ws, widths)


def write_schedule_i_detail(wb: Workbook, conn) -> None:
    """Schedule I / Part XV grant-line detail — every non-unrelated grant.
    This is the auditable filing-verified record."""
    rows = conn.execute(
        """
        SELECT df.donor_ticker, df.foundation_name, g.tax_year,
               g.grantee_classification, g.recipient_name, g.recipient_ein,
               g.grant_amount, g.grant_purpose, g.source_filing_url
        FROM foundation_grants g
        JOIN donor_foundations df ON df.ein = g.donor_ein
        WHERE g.source = 'irs_xml' AND g.grantee_classification != 'unrelated'
        ORDER BY
            CASE g.grantee_classification
                WHEN 'anti_police_adversarial'       THEN 1
                WHEN 'pro_police'                    THEN 2
                WHEN 'collaborative_reform'          THEN 3
                WHEN 'broad_cj_reform'               THEN 4
                WHEN 'reentry_employment'            THEN 5
                WHEN 'innocence_wrongful_conviction' THEN 6
                ELSE 9 END,
            g.grant_amount DESC;
        """
    ).fetchall()
    if not rows:
        return

    ws = wb.create_sheet("Schedule I detail (990)", 3)
    headers = ["Donor ticker", "Donor foundation", "Tax year", "Classification",
               "Recipient", "Recipient EIN", "Grant amount", "Purpose", "Source filing"]
    ws.append(headers)
    _style_header(ws, len(headers))

    classification_fills = {
        "anti_police_adversarial":       PatternFill("solid", fgColor="DCE7F2"),
        "collaborative_reform":          PatternFill("solid", fgColor="E2EFDA"),
        "broad_cj_reform":               PatternFill("solid", fgColor="FFF2CC"),
        "reentry_employment":            PatternFill("solid", fgColor="EDEDED"),
        "innocence_wrongful_conviction": PatternFill("solid", fgColor="FCE4D6"),
        "pro_police":                    PatternFill("solid", fgColor="F4CCCC"),
    }
    for r in rows:
        ws.append([
            r["donor_ticker"],
            r["foundation_name"],
            r["tax_year"],
            r["grantee_classification"],
            r["recipient_name"],
            r["recipient_ein"] or "",
            r["grant_amount"] or 0,
            r["grant_purpose"] or "",
            r["source_filing_url"] or "",
        ])
        row_idx = ws.max_row
        fill = classification_fills.get(r["grantee_classification"], PatternFill())
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = fill
            cell.alignment = WRAP
        ws.cell(row=row_idx, column=7).number_format = '"$"#,##0'
        ws.row_dimensions[row_idx].height = 50

    _set_widths(ws, {1: 12, 2: 36, 3: 9, 4: 30, 5: 50, 6: 14, 7: 14, 8: 55, 9: 50})


def write_990_breakdown(wb: Workbook, conn) -> None:
    """Pivot: foundation × classification with grant counts + totals."""
    rows = conn.execute(
        """
        SELECT df.donor_ticker, df.foundation_name,
               g.grantee_classification,
               COUNT(*) AS n_grants,
               COALESCE(SUM(g.grant_amount), 0) AS total_amount
        FROM foundation_grants g
        JOIN donor_foundations df ON df.ein = g.donor_ein
        WHERE g.source = 'irs_xml'
        GROUP BY df.donor_ticker, g.grantee_classification
        ORDER BY df.donor_ticker, total_amount DESC;
        """
    ).fetchall()
    if not rows:
        return

    ws = wb.create_sheet("990 Classification Breakdown", 4)
    headers = ["Donor ticker", "Foundation", "Classification", "# grants", "Total $"]
    ws.append(headers)
    _style_header(ws, len(headers))

    classification_fills = {
        "anti_police_adversarial":       PatternFill("solid", fgColor="DCE7F2"),
        "collaborative_reform":          PatternFill("solid", fgColor="E2EFDA"),
        "broad_cj_reform":               PatternFill("solid", fgColor="FFF2CC"),
        "reentry_employment":            PatternFill("solid", fgColor="EDEDED"),
        "innocence_wrongful_conviction": PatternFill("solid", fgColor="FCE4D6"),
        "pro_police":                    PatternFill("solid", fgColor="F4CCCC"),
        "unrelated":                     PatternFill("solid", fgColor="F2F2F2"),
    }
    for r in rows:
        ws.append([
            r["donor_ticker"], r["foundation_name"],
            r["grantee_classification"] or "(none)", r["n_grants"], r["total_amount"],
        ])
        row_idx = ws.max_row
        fill = classification_fills.get(r["grantee_classification"], PatternFill())
        for c in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=c).fill = fill
        ws.cell(row=row_idx, column=5).number_format = '"$"#,##0'

    _set_widths(ws, {1: 12, 2: 36, 3: 32, 4: 10, 5: 16})


def write_federal_le_contracts(wb: Workbook, conn) -> None:
    """All federal LE-agency contracts. The hard-numbers verification of pro-police claims."""
    rows = conn.execute(
        """
        SELECT cfc.ticker, csi.net_position, cfc.recipient_name,
               cfc.awarding_agency, cfc.awarding_sub_agency,
               cfc.period_start, cfc.period_end, cfc.award_amount,
               cfc.description, cfc.award_id
        FROM company_federal_contracts cfc
        LEFT JOIN company_stance_investigation csi ON csi.ticker = cfc.ticker
        WHERE cfc.is_le_agency = 1
        ORDER BY cfc.award_amount DESC;
        """
    ).fetchall()
    if not rows:
        return
    ws = wb.create_sheet("Federal LE Contracts", 4)
    headers = ["Ticker", "Net Position", "Awardee", "Awarding Agency",
               "Sub-Agency", "Period start", "Period end", "Award Amount",
               "Description", "Award ID"]
    ws.append(headers)
    _style_header(ws, len(headers))
    for r in rows:
        net = r["net_position"] or "unknown"
        ws.append([
            r["ticker"], net, r["recipient_name"],
            r["awarding_agency"], r["awarding_sub_agency"],
            r["period_start"], r["period_end"],
            r["award_amount"] or 0,
            (r["description"] or "")[:300],
            r["award_id"],
        ])
        row_idx = ws.max_row
        fill = NET_FILLS.get(net, PatternFill())
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = fill
            cell.alignment = WRAP
        ws.cell(row=row_idx, column=8).number_format = '"$"#,##0'
        ws.row_dimensions[row_idx].height = 40
    _set_widths(ws, {1: 7, 2: 18, 3: 32, 4: 30, 5: 38,
                     6: 12, 7: 12, 8: 16, 9: 60, 10: 22})


def write_daf_anti_adjacent(wb: Workbook, conn) -> None:
    """Every DAF grant to a police-adjacent recipient (the black-box that we cracked open)."""
    rows = conn.execute(
        """
        SELECT df.foundation_name, df.ein, g.tax_year, g.grantee_classification,
               g.recipient_name, g.recipient_ein, g.grant_amount, g.grant_purpose
        FROM foundation_grants g
        JOIN donor_foundations df ON df.ein = g.donor_ein
        WHERE df.relationship = 'daf'
          AND g.grantee_classification NOT IN ('unrelated', 'daf_intermediary')
        ORDER BY
          CASE g.grantee_classification
              WHEN 'anti_police_adversarial'       THEN 1
              WHEN 'pro_police'                    THEN 2
              WHEN 'collaborative_reform'          THEN 3
              WHEN 'broad_cj_reform'               THEN 4
              WHEN 'reentry_employment'            THEN 5
              WHEN 'innocence_wrongful_conviction' THEN 6
              ELSE 9 END,
          g.grant_amount DESC;
        """
    ).fetchall()
    if not rows:
        return
    ws = wb.create_sheet("DAF Outflows (Schedule I)", 5)
    headers = ["DAF Foundation", "DAF EIN", "Tax Year", "Classification",
               "Recipient", "Recipient EIN", "Grant Amount", "Purpose"]
    ws.append(headers)
    _style_header(ws, len(headers))

    classification_fills = {
        "anti_police_adversarial":       PatternFill("solid", fgColor="DCE7F2"),
        "collaborative_reform":          PatternFill("solid", fgColor="E2EFDA"),
        "broad_cj_reform":               PatternFill("solid", fgColor="FFF2CC"),
        "reentry_employment":            PatternFill("solid", fgColor="EDEDED"),
        "innocence_wrongful_conviction": PatternFill("solid", fgColor="FCE4D6"),
        "pro_police":                    PatternFill("solid", fgColor="F4CCCC"),
    }
    for r in rows:
        ws.append([
            r["foundation_name"], r["ein"], r["tax_year"],
            r["grantee_classification"], r["recipient_name"],
            r["recipient_ein"] or "", r["grant_amount"] or 0,
            r["grant_purpose"] or "",
        ])
        row_idx = ws.max_row
        fill = classification_fills.get(r["grantee_classification"], PatternFill())
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = fill
            cell.alignment = WRAP
        ws.cell(row=row_idx, column=7).number_format = '"$"#,##0'
        ws.row_dimensions[row_idx].height = 40
    _set_widths(ws, {1: 36, 2: 11, 3: 9, 4: 32, 5: 42, 6: 11, 7: 14, 8: 55})


def write_police_foundation_funders(wb: Workbook, rows: list[dict]) -> None:
    """Focused list: the donates_to_police_foundations subset."""
    ws = wb.create_sheet("Police-foundation funders")
    headers = [
        "Ticker", "Company", "Sector", "First donation date", "Current status",
        "Net position", "Summary", "Evidence URL",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    filtered = [r for r in rows if r["pro_police_type"] == "donates_to_police_foundations"]
    filtered.sort(key=lambda r: (r["pro_police_first_date"] or "9999", r["ticker"]))

    for r in filtered:
        net = r["net_position"] or "unknown"
        fill = NET_FILLS.get(net, PatternFill())
        ws.append([
            r["ticker"],
            r["company_name"],
            r["sector"],
            r["pro_police_first_date"] or "",
            r["pro_police_current_status"] or "",
            net,
            r["pro_police_summary"] or "",
            "",
        ])
        row_idx = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = fill
            cell.alignment = WRAP
            cell.font = DEFAULT_FONT
        if r["pro_police_evidence_url"]:
            _link(ws.cell(row=row_idx, column=8), r["pro_police_evidence_url"], "source")
        ws.row_dimensions[row_idx].height = 100

    _set_widths(ws, {1: 8, 2: 28, 3: 22, 4: 16, 5: 16, 6: 18, 7: 75, 8: 12})


def main() -> None:
    conn = connect(RISK_DB_PATH)
    rows = fetch_rows(conn)
    print(f"Loaded {len(rows)} investigation rows.")

    wb = Workbook()
    # Drop default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Build sheets
    write_summary(wb, rows)
    write_990_trajectory(wb, conn)
    write_990_breakdown(wb, conn)
    write_schedule_i_detail(wb, conn)
    write_federal_le_contracts(wb, conn)
    write_daf_anti_adjacent(wb, conn)
    write_master(wb, rows, sheet_name="All Companies")
    write_master(wb, rows, sheet_name="Reform-leaning",
                 filter_fn=lambda r: r["net_position"] == "reform_leaning")
    write_master(wb, rows, sheet_name="Cross-exposure",
                 filter_fn=lambda r: r["net_position"] == "cross_exposure")
    write_master(wb, rows, sheet_name="Enforcement-leaning",
                 filter_fn=lambda r: r["net_position"] == "enforcement_leaning")
    write_focused_anti(wb, rows)
    write_focused_pro(wb, rows)
    write_police_foundation_funders(wb, rows)

    stamp = datetime.now(timezone.utc).strftime("%Y%m%d")
    out_path = Path(DATA_DIR) / f"police_policy_stance_{stamp}.xlsx"
    wb.save(out_path)
    print(f"Wrote -> {out_path}")
    print(f"  {len(wb.sheetnames)} sheets: {wb.sheetnames}")

    conn.close()


if __name__ == "__main__":
    main()
