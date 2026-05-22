"""
Add a "Sources & Methodology" sheet to the Excel workbook listing every data
source used in this investigation. Goes after the Summary sheet so it's
visible up-front when the client opens the workbook.
"""
from __future__ import annotations

import sys
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config import DATA_DIR, RISK_DB_PATH
from src.risk_schema import connect

HEADER_FILL = PatternFill("solid", fgColor="333333")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True)
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="333333")
SECTION_FILL = PatternFill("solid", fgColor="DDDDDD")
LINK_FONT = Font(name="Calibri", size=10, color="0563C1", underline="single")
WRAP = Alignment(wrap_text=True, vertical="top")


# All sources used in the investigation, in order of importance to the final dataset.
SOURCES = [
    # --- Primary structured data ---
    {
        "category": "Financial / Market data",
        "name": "Financial Modeling Prep (FMP)",
        "type": "Commercial REST API",
        "provided": "S&P 500 constituents (503 tickers with sector); per-ticker news/press-release feeds; general market news (3,909 articles ingested into the `articles` table)",
        "url": "https://financialmodelingprep.com/stable/",
        "cost": "Paid subscription (Ultimate plan)",
        "tables_populated": "articles, article_keyword_hits, companies",
        "row_count": "3,909 articles; 503 universe tickers",
        "limitations": "Finance/market-focused corpus. ZERO hits on police-specific keywords across 3,909 articles — wrong corpus for political/social events. No Russell 1000 endpoint (used S&P 500 as proxy).",
        "license": "FMP Ultimate Plan",
    },
    {
        "category": "AI investigation",
        "name": "Anthropic Claude API (Opus 4.7) with server-side web_search",
        "type": "LLM API with built-in web search tool",
        "provided": "Per-company investigation: reform-side and enforcement-side actions, dates, current status, evidence URLs, confidence; v3 taxonomy classification",
        "url": "https://api.anthropic.com  (web_search_20260209)",
        "cost": "Paid per-token; ~$25-50 for full sweep of 254 tickers",
        "tables_populated": "company_stance_investigation",
        "row_count": "254 company investigations",
        "limitations": "AI judgment with confidence scores; primary-source URLs returned but require validation. Claude's web_search uses Anthropic's own search index, not direct Google.",
        "license": "Anthropic API terms",
    },
    {
        "category": "Nonprofit / charitable giving",
        "name": "IRS Form 990 / 990-PF XML",
        "type": "Direct IRS XML monthly archives",
        "provided": "Foundation filings: annual total grants paid + Schedule I / Part XV line-item grant detail (recipient name, EIN, amount, purpose)",
        "url": "https://apps.irs.gov/pub/epostcard/990/xml/{2023,2024}/  (monthly TEOS_XML ZIPs) + https://apps.irs.gov/pub/epostcard/990/xml/{year}/index_{year}.csv",
        "cost": "Free",
        "tables_populated": "foundation_filings, foundation_grants (irs_xml source)",
        "row_count": "186 foundation filings parsed; 29,443 grant-line records (corporate foundations + DAFs)",
        "limitations": "FY2022 most-recent fully covered; some FY2023 available. 8.1 GB of raw XML zips on disk. Schedule B (donor names) is publicly redacted on most filings.",
        "license": "Public domain (US Federal government data)",
    },
    {
        "category": "Nonprofit / charitable giving",
        "name": "ProPublica Nonprofit Explorer API",
        "type": "Free REST API",
        "provided": "EIN search by org name; structured filing metadata; total contributions paid trajectory (used to find donor foundation EINs, validate filings exist before pulling raw XML)",
        "url": "https://projects.propublica.org/nonprofits/api/v2/  (search.json, organizations/{ein}.json)",
        "cost": "Free, no API key",
        "tables_populated": "donor_foundations",
        "row_count": "30 donor foundations mapped (21 corporate + 9 DAFs)",
        "limitations": "Top-level 990 fields only — does NOT expose Schedule I line items via API (had to parse raw IRS XML for those).",
        "license": "Open data — ProPublica terms",
    },
    {
        "category": "Federal contracting",
        "name": "USAspending.gov",
        "type": "Federal government REST API",
        "provided": "Federal contract awards by recipient: awarding agency, sub-agency (FBI/DEA/ATF/USMS/ICE/CBP/etc.), period of performance, award amount, description",
        "url": "https://api.usaspending.gov/api/v2/search/spending_by_award/",
        "cost": "Free, no API key",
        "tables_populated": "company_federal_contracts",
        "row_count": "8,048 federal contracts ingested across 94 tickers (FY2020-FY2025); 2,524 flagged as law-enforcement-agency awards totaling $11B+",
        "limitations": "Recipient-name fuzzy matching; subsidiary contracts may attribute to parent. Coverage 5 ticker-level connection errors (LDOS, LEN, LH, LVS, MGM, MSI) — worth retry pass.",
        "license": "Public domain (US Federal government data)",
    },
    {
        "category": "Corporate disclosure",
        "name": "SEC EDGAR full-text search",
        "type": "Free REST API",
        "provided": "Filings matching keyword search per CIK: 10-K, 10-Q, DEF 14A (proxy statements), 8-K. Keywords: 'defund police', 'EEOC investigation', 'FCC license review', 'police reform', 'racial equity', 'Black Lives Matter', 'criminal justice reform'.",
        "url": "https://efts.sec.gov/LATEST/search-index  +  https://www.sec.gov/files/company_tickers.json",
        "cost": "Free (must respect 10 req/sec rate limit)",
        "tables_populated": "company_sec_signals",
        "row_count": "135 ticker-specific keyword hits across 94 tickers (CIK-filtered)",
        "limitations": "Foreign-listed issuers (UL/Unilever) file 20-F not 10-K, so EDGAR misses them. Initial run used ticker filter (broken) — corrected to CIK filter.",
        "license": "Public domain (SEC.gov terms)",
    },
    # --- Article text extraction ---
    {
        "category": "Full-text article extraction",
        "name": "trafilatura (Python library)",
        "type": "HTML-to-article extractor",
        "provided": "Full body text from news article URLs (clean-stripped of nav/ads/footers); used for full-text scoring v2",
        "url": "https://github.com/adbar/trafilatura  (PyPI install)",
        "cost": "Free / open-source MIT",
        "tables_populated": "article_text",
        "row_count": "5 origin + 1 wayback successes out of 9 attempts (the rest behind paywalls)",
        "limitations": "Defeated by WSJ/Reuters/Fast Company paywalls (HTTP 401/403). YouTube URLs unsupported.",
        "license": "MIT (open source)",
    },
    {
        "category": "Full-text article extraction",
        "name": "Internet Archive Wayback Machine",
        "type": "Free availability API",
        "provided": "Fallback URL when origin returns 401/403/short. Recovered 1 of 3 paywalled WSJ pieces in our test set.",
        "url": "https://archive.org/wayback/available + http://web.archive.org/web/...",
        "cost": "Free",
        "tables_populated": "article_text (fetch_source='wayback')",
        "row_count": "1 paywalled article recovered (WSJ Disney FCC piece)",
        "limitations": "Reuters and Fast Company actively block archive.org — no snapshots available for those domains.",
        "license": "Public benefit",
    },
    # --- Attempted but failed ---
    {
        "category": "Global news (attempted)",
        "name": "GDELT 2.0 DOC API",
        "type": "Free REST API",
        "provided": "Intended: police-related news across 100+ languages for 94 watchlist tickers 2020-2026",
        "url": "https://api.gdeltproject.org/api/v2/doc/doc",
        "cost": "Free, no API key",
        "tables_populated": "gdelt_news (currently empty)",
        "row_count": "0 — API timed out / rate-limited on most queries",
        "limitations": "API throttling killed the run mid-way. Retry on a different network or use alternative news API (NewsAPI, Event Registry, Common Crawl News).",
        "license": "GDELT terms",
    },
    # --- Curated inputs / methodology ---
    {
        "category": "Curated inputs (our own)",
        "name": "Manual watchlist seeds",
        "type": "Curated CSV files",
        "provided": "Initial company watchlist (data/company_master.csv), keyword ontology (data/query_taxonomy.csv with 23 reform/enforcement/governance keywords), manually-curated known anti/pro stances (data/anti_police_companies.csv)",
        "url": "Internal — see data/ directory in project",
        "cost": "Analyst time",
        "tables_populated": "companies, query_taxonomy",
        "row_count": "3 companies seed; 23 keyword ontology entries; 7 manual stance entries",
        "limitations": "Subjective curation; v3 refined taxonomy after analyst review identified misclassification (e.g. Chubb Rule of Law = collaborative reform, not anti-police).",
        "license": "Internal",
    },
    {
        "category": "Curated inputs (our own)",
        "name": "Donor foundation EIN mapping",
        "type": "Curated CSV",
        "provided": "Ticker → corporate-foundation EIN mapping (21 corporate foundations + 9 DAF intermediaries). Built via ProPublica search + manual research.",
        "url": "data/donor_foundations.csv",
        "cost": "Analyst time",
        "tables_populated": "donor_foundations",
        "row_count": "30 EINs",
        "limitations": "Some tickers (LULU, MSFT, AMZN US, HD) have no separate 501c3 foundation — give directly from corporate entity (no 990 filed for those flows).",
        "license": "Internal",
    },
    # --- Methodology / classification ---
    {
        "category": "Methodology",
        "name": "Policy-stance scoring rule-engine (v4)",
        "type": "Deterministic Python rules",
        "provided": "Signed -5.0 to +5.0 score per company derived from: reform-type weight + enforcement-type weight + status modifier + federal LE contract bonus. Sign: + = enforcement-leaning, − = reform-leaning.",
        "url": "scripts/apply_option_a_and_score.py",
        "cost": "—",
        "tables_populated": "company_stance_investigation.policy_stance_score",
        "row_count": "254 scores computed",
        "limitations": "Weights are analyst-set (REFORM_WEIGHTS, ENFORCE_WEIGHTS dicts) and tunable. Mixed companies (cross_exposure) net toward zero by design.",
        "license": "Internal",
    },
    {
        "category": "Methodology",
        "name": "Grantee classifier (v3 taxonomy)",
        "type": "Rule-based regex matcher",
        "provided": "Classifies 990 Schedule I grant recipients into: reform_advocacy_grants (BLM/ArchCity/Min Freedom Fund), collaborative_reform (Policing Project @ NYU), broad_cj_reform (NAACP LDF/EJI/ACLU), reentry_employment (Anti-Recidivism Coalition/Defy Ventures), innocence_wrongful_conviction (Innocence Project family), pro_police (Police Foundation/PBA/FOP), daf_intermediary (Tides/AOGF/NPT/Fidelity Charitable/etc.), unrelated.",
        "url": "scripts/classify_grants.py",
        "cost": "—",
        "tables_populated": "foundation_grants.grantee_classification",
        "row_count": "29,443 grants classified",
        "limitations": "Pattern matching only — relies on canonical recipient names. Edge cases require manual review (e.g. ACLU classified as broad_cj_reform even though their specific programs vary).",
        "license": "Internal",
    },
]


def write_sources_sheet(wb, conn) -> None:
    # Find a position right after Summary
    if "Sources & Methodology" in wb.sheetnames:
        del wb["Sources & Methodology"]
    ws = wb.create_sheet("Sources & Methodology", 1)

    # Title
    ws["A1"] = "Sources & Methodology"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = (
        "Every data source used in this investigation, what it provided, "
        "cost, and known limitations. Generated "
        + datetime.now(timezone.utc).strftime("%Y-%m-%d")
    )
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:F2")
    ws.row_dimensions[1].height = 22

    # Headers
    headers = ["#", "Source", "Type", "What it provided", "URL / endpoint",
               "Cost", "Tables populated", "Row count", "Limitations", "License"]
    ws.append([])  # blank row
    ws.append(headers)
    header_row = ws.max_row
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 30
    ws.freeze_panes = f"A{header_row+1}"

    last_category = None
    n = 0
    for src in SOURCES:
        # Section divider on category change
        if src["category"] != last_category:
            ws.append([src["category"]])
            sec_row = ws.max_row
            sec_cell = ws.cell(row=sec_row, column=1)
            sec_cell.font = SECTION_FONT
            sec_cell.fill = SECTION_FILL
            ws.merge_cells(start_row=sec_row, start_column=1, end_row=sec_row, end_column=len(headers))
            ws.cell(row=sec_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[sec_row].height = 22
            last_category = src["category"]

        n += 1
        ws.append([
            n,
            src["name"],
            src["type"],
            src["provided"],
            src["url"],
            src["cost"],
            src["tables_populated"],
            src["row_count"],
            src["limitations"],
            src["license"],
        ])
        row_idx = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.alignment = WRAP
            cell.font = Font(name="Calibri", size=10)

        # Hyperlink the URL cell if it starts with http
        url = src["url"]
        if isinstance(url, str) and url.startswith("http"):
            url_cell = ws.cell(row=row_idx, column=5)
            # Take first URL only if multi-line
            first_url = url.split()[0]
            url_cell.hyperlink = first_url
            url_cell.font = LINK_FONT

        ws.row_dimensions[row_idx].height = 110

    # Column widths
    widths = {1: 4, 2: 32, 3: 26, 4: 60, 5: 50, 6: 24, 7: 32, 8: 24, 9: 60, 10: 22}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Footer / methodology note
    ws.append([])
    ws.append([])
    ws.append(["Investigation phases (sequence the data flowed through)"])
    ws.cell(row=ws.max_row, column=1).font = SECTION_FONT
    ws.cell(row=ws.max_row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len(headers))

    phases = [
        ("Phase 0", "FMP news scrape + keyword-match against query_taxonomy — confirmed FMP corpus has zero police-specific hits across 3,909 articles."),
        ("Phase 1", "Claude (Opus 4.7 + web_search) investigates each ticker → company_stance_investigation. Started with 3 watchlist tickers, expanded to S&P 500 priority sectors (254 tickers)."),
        ("Phase 2", "Taxonomy refinement (v3): split 'donations_police_reform' into adversarial / collaborative / re-entry / innocence sub-types. Re-investigated 30 anti-flagged tickers under refined taxonomy — 18 moved out of 'anti'."),
        ("Phase 3", "IRS Form 990 verification: located 21 donor foundations + 9 DAFs in IRS index, downloaded 24 monthly XML ZIPs (8.1 GB), parsed Schedule I/Part XV → 29,443 grant-line records classified."),
        ("Phase 4A", "DAF outflow scan: parsed 21 DAF filings, found $1.99M flowing through DAFs to adversarial-anti-police orgs (BLM, Critical Resistance, ArchCity Defenders) — invisible from corporate foundation 990s alone."),
        ("Phase 4B", "USAspending federal-contracting layer: $11B+ in law-enforcement-agency contracts verified across 94 tickers — hardens 'enforcement-leaning' classification with primary-source dollar amounts."),
        ("Phase 4C", "SEC EDGAR keyword scan (CIK-filtered): 135 ticker-specific filings mentioning police-reform / racial-equity / EEOC / FCC keywords 2020-2026."),
        ("Phase 5", "Option A relabeling (reform_leaning / enforcement_leaning / cross_exposure / no_material_exposure) + signed -5.0..+5.0 policy_stance_score computed from existing data."),
    ]
    for phase, desc in phases:
        ws.append([phase, desc])
        row_idx = ws.max_row
        ws.cell(row=row_idx, column=1).font = Font(bold=True)
        ws.cell(row=row_idx, column=2).alignment = WRAP
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=len(headers))
        ws.row_dimensions[row_idx].height = 50


def main() -> None:
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d")
    xlsx_path = Path(DATA_DIR) / f"police_policy_stance_{stamp}.xlsx"
    if not xlsx_path.exists():
        print(f"Excel workbook not found: {xlsx_path}")
        return

    wb = load_workbook(xlsx_path)
    conn = connect(RISK_DB_PATH)
    write_sources_sheet(wb, conn)
    wb.save(xlsx_path)
    print(f"Added Sources & Methodology sheet -> {xlsx_path}")
    print(f"Workbook now has {len(wb.sheetnames)} sheets")
    conn.close()


if __name__ == "__main__":
    main()
